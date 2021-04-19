#Versão IAM

#import yfinance as yf
from yahooquery import Ticker
#import yahooquery as yq
import openpyxl as xl
import pandas as pd
import numpy as np
#from openpyxl.chart import BarChart, Reference

#ativo = yf.Ticker("GNDI3.SA")
#print(ativo)
"""
returns
<yfinance.Ticker object at 0x1a1715e898>
"""

# get stock info
#print(ativo.info)

"""
returns:
{
 'quoteType': 'EQUITY',
 'quoteSourceName': 'Nasdaq Real Time Price',
 'currency': 'USD',
 'shortName': 'Microsoft Corporation',
 'exchangeTimezoneName': 'America/New_York',
  ...
 'symbol': 'MSFT'
}
"""

# get historical market data, here max is 5 years.
#ativo.history(period="max")

#print(ativo.history())
mf_symbols = pd.read_excel('C:\\Users\\juliano.pulicci\\Documents\\Pessoais\\Trader\\Distorcao\\Yahoo.xlsx',sheet_name='Tickers')
simbolos = mf_symbols.values.tolist()
# Tickers da Planilha Yahoo.xlsx


# Add symbol column to each dataframe
#for key in data:
#    if isinstance(data[key], pd.DataFrame):
#        data[key]['symbol'] = key
# Concat all dataframes
#df = pd.concat([data[k] for k in data if isinstance(data[k], pd.DataFrame)])

# Datas específicas
#petr.history(start='2005-05-01', end='2021-03-30',  interval = "1d")

# Intraday - 30 minutos
abev = Ticker('ABEV3.SA')
abev.history(period='60d',  interval = "30m")

# Intraday - 1 minuto
#abev = abev.history(period='7d',  interval = "1m")

# Informações financeiras
#petr = Ticker("PETR4.SA")     # Coleta dados
#petr = petr.income_statement()# Chama função de Demonstração de resultados
#petr = petr.transpose()       # Transpõe a matriz
#petr.columns = petr.iloc[0,:] # Renomeia colunas
#petr = petr.iloc[2:,:-1]      # Seleciona dados
#petr = petr.iloc[:, ::-1]     # Inverte colunas

wb = xl.load_workbook('C:\\Users\\juliano.pulicci\\Documents\\Pessoais\\Trader\\Distorcao\\Yahoo.xlsx')
sheet = wb['Dados']
lin=2
for item in simbolos :
    petr = Ticker(item, asynchronous=True)
    obj = petr.history(period='2d',  interval = "1d")
    #obj = obj.reset_index(level=[0,1])
    Tipo = str(type(obj))
    if str(Tipo) <> 'dict':
        lin = len(obj.index)
        col = len(obj.columns)
        #lin=len(obj.count())
        #if isinstance(obj.columns, list):
        for row in range(lin):
            for colu in range(col):
                valor = obj.iat[row,colu]
                sheet.cell(lin, colu+1).value = valor
    lin=lin+1
#    cell = sheet.cell(row, 3)
#    corrected_price = float(cell.value.replace('$', '')) * 0.9
#    corrected_price_cell = sheet.cell(row, 4)
#    corrected_price_cell.value = corrected_price

#values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
#chart = BarChart()
#chart.add_data(values)
#sheet.add_chart(chart, 'e2')

wb.save('C:\\Users\\juliano.pulicci\\Documents\\Pessoais\\Trader\\Distorcao\\Yahoo.xlsx')
